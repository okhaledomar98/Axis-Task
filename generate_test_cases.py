from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

# ── Colors ──────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
ORANGE      = "FF8C00"
GREEN       = "1E7145"
RED         = "C00000"

# ── Helper: make a Fill ──────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(color=WHITE, size=11):
    return Font(bold=True, color=color, size=size)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_center():
    return Alignment(wrap_text=True, vertical="top", horizontal="center")

def wrap_left():
    return Alignment(wrap_text=True, vertical="top", horizontal="left")

# ── Title Row ────────────────────────────────────────────
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "Tealium Ecommerce Demo — Test Cases"
title_cell.fill = fill(DARK_BLUE)
title_cell.font = Font(bold=True, color=WHITE, size=16)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Sub-title Row ────────────────────────────────────────
ws.merge_cells("A2:H2")
sub_cell = ws["A2"]
sub_cell.value = "Features Covered: User Registration | User Login | Add to Cart (E2E)"
sub_cell.fill = fill(MID_BLUE)
sub_cell.font = Font(bold=False, color=WHITE, size=11, italic=True)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# ── Header Row ───────────────────────────────────────────
headers = [
    "TC ID", "Feature", "Test Case Description",
    "Pre-conditions", "Test Steps", "Test Data",
    "Expected Result", "Status"
]
col_widths = [8, 14, 32, 28, 42, 28, 32, 10]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.fill = fill(MID_BLUE)
    cell.font = bold_font(WHITE, 11)
    cell.alignment = wrap_center()
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[3].height = 22

# ── Test Case Data ───────────────────────────────────────
test_cases = [
    {
        "id": "TC-01",
        "feature": "User Registration",
        "description": "Verify a new user can register successfully with all valid inputs",
        "preconditions": (
            "1. Browser is open\n"
            "2. Site is accessible\n"
            "3. User is NOT already registered with this email"
        ),
        "steps": (
            "1. Navigate to https://ecommerce.tealiumdemo.com\n"
            "2. Click on the Account icon\n"
            "3. Click 'Register'\n"
            "4. Fill in First Name\n"
            "5. Fill in Last Name\n"
            "6. Fill in Email address\n"
            "7. Fill in Password\n"
            "8. Confirm Password\n"
            "9. Click 'Register' button"
        ),
        "data": (
            "First Name: Test\n"
            "Last Name: User\n"
            "Email: axis.testuser@gmail.com\n"
            "Password: Adriano@123\n"
            "Confirm: Adriano@123"
        ),
        "expected": (
            "User is redirected to Account Dashboard.\n"
            "Welcome message displays the registered user's name.\n"
            "No error messages are shown."
        ),
        "status": "PASS"
    },
    {
        "id": "TC-02",
        "feature": "User Registration",
        "description": "Verify that registration fails when required fields are left empty",
        "preconditions": (
            "1. Browser is open\n"
            "2. Site is accessible\n"
            "3. User is on the Registration page"
        ),
        "steps": (
            "1. Navigate to the Register page\n"
            "2. Leave all fields empty\n"
            "3. Click 'Register' button"
        ),
        "data": (
            "All fields: empty"
        ),
        "expected": (
            "User remains on the Registration page.\n"
            "Required field validation errors appear on all empty fields.\n"
            "Registration does NOT proceed."
        ),
        "status": "PASS"
    },
    {
        "id": "TC-03",
        "feature": "User Login",
        "description": "Verify a registered user can log in successfully with valid credentials",
        "preconditions": (
            "1. Browser is open\n"
            "2. Site is accessible\n"
            "3. A registered account exists with the test credentials"
        ),
        "steps": (
            "1. Navigate to https://ecommerce.tealiumdemo.com\n"
            "2. Click on the Account icon\n"
            "3. Click 'Log In'\n"
            "4. Enter Email address\n"
            "5. Enter Password\n"
            "6. Click 'Login' button"
        ),
        "data": (
            "Email: o.khaledomar98@gmail.com\n"
            "Password: 1234567"
        ),
        "expected": (
            "User is redirected to Account Dashboard.\n"
            "Welcome message shows 'Hello, Omar Khaled!'.\n"
            "No error messages are shown."
        ),
        "status": "PASS"
    },
    {
        "id": "TC-04",
        "feature": "User Login",
        "description": "Verify that login fails when incorrect password is entered",
        "preconditions": (
            "1. Browser is open\n"
            "2. Site is accessible\n"
            "3. A registered account exists with the test email"
        ),
        "steps": (
            "1. Navigate to the Login page\n"
            "2. Enter a valid registered email\n"
            "3. Enter a wrong password\n"
            "4. Click 'Login' button"
        ),
        "data": (
            "Email: o.khaledomar98@gmail.com\n"
            "Password: WrongPassword999!"
        ),
        "expected": (
            "User remains on the Login page.\n"
            "Error message is displayed: 'Invalid login or password.'\n"
            "User is NOT redirected to the dashboard."
        ),
        "status": "PASS"
    },
    {
        "id": "TC-05",
        "feature": "Add to Cart (E2E)",
        "description": (
            "End-to-end flow: Login → Navigate to Shoes via Accessories menu → "
            "Sort by Price → Select Dorian product → Choose Color & Size → Add to Cart"
        ),
        "preconditions": (
            "1. Browser is open\n"
            "2. Site is accessible\n"
            "3. A registered account exists\n"
            "4. 'Dorian Perforated Oxford' product is available in Shoes category"
        ),
        "steps": (
            "1. Login with valid credentials\n"
            "2. Hover over 'Accessories' in the navigation menu\n"
            "3. Verify the Accessories dropdown appears\n"
            "4. Click 'Shoes' from the dropdown\n"
            "5. Verify Shoes page is loaded\n"
            "6. Sort products by 'Price' (Low to High)\n"
            "7. Verify prices are in ascending order\n"
            "8. Click on 'Dorian Perforated Oxford' product\n"
            "9. Verify product details page is loaded\n"
            "10. Select first available Color\n"
            "11. Verify color is selected\n"
            "12. Select first available Size\n"
            "13. Verify size is selected\n"
            "14. Click 'Add to Cart'"
        ),
        "data": (
            "Email: o.khaledomar98@gmail.com\n"
            "Password: 1234567\n"
            "Product: Dorian Perforated Oxford\n"
            "Color: Black (first available)\n"
            "Size: first available"
        ),
        "expected": (
            "User is redirected to Shopping Cart page.\n"
            "A success message is displayed confirming the product was added.\n"
            "Cart icon shows updated item count."
        ),
        "status": "PASS"
    },
]

# ── Write Test Case Rows ─────────────────────────────────
for tc_idx, tc in enumerate(test_cases):
    row = 4 + tc_idx
    row_fill = fill(WHITE) if tc_idx % 2 == 0 else fill(LIGHT_GRAY)

    values = [
        tc["id"], tc["feature"], tc["description"],
        tc["preconditions"], tc["steps"], tc["data"],
        tc["expected"], tc["status"]
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = row_fill
        cell.border = thin_border()

        if col_idx == 1:
            cell.font = Font(bold=True, color=MID_BLUE, size=10)
            cell.alignment = wrap_center()
        elif col_idx == 2:
            cell.font = Font(bold=True, color=DARK_BLUE, size=10)
            cell.alignment = wrap_center()
        elif col_idx == 8:
            # Status coloring
            if value == "PASS":
                cell.fill = fill("C6EFCE")
                cell.font = Font(bold=True, color=GREEN, size=10)
            else:
                cell.fill = fill("FFC7CE")
                cell.font = Font(bold=True, color=RED, size=10)
            cell.alignment = wrap_center()
        else:
            cell.font = Font(size=10)
            cell.alignment = wrap_left()

    ws.row_dimensions[row].height = 110

# ── Freeze header rows ───────────────────────────────────
ws.freeze_panes = "A4"

# ── Save ─────────────────────────────────────────────────
output_path = r"D:\Automation Tasks\Tasks\tealium-automation-framework\Test_Cases_Tealium_Ecommerce.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
